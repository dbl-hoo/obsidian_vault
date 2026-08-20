"""Insert or update a candidate row in a Project Mercury MSA survey workbook.

Usage:
    python add_candidate.py --msa Cincinnati --type building --json candidate.json [--pin Cincinnati41018] [--overwrite]

The workbook is self-describing: pin band rows (grey rows with the pin ID in
col B and "... pin (lat, lon)" in col C) define the groups. The script assigns
the candidate to the nearest pin (unless --pin forces one), dedupes against
existing rows by address/name, keeps each group sorted by distance, renumbers,
and appends criteria flags to Comments.

JSON fields (all optional except name-or-address):
  common: name, address, city, state, zip, lat, lon, owner, flyer, comments,
          sale_lease, sale_price, power
  building: total_sf, avail_sf, office_sf, land_ac, year_built, building_status,
            building_type, occupancy, tenancy, current_tenant, prior_use,
            prior_tenant, date_available, sprinkler, zoning, clear_min, clear_max,
            load_type, docks, grade_doors, col_w, col_d, auto_parking,
            trailer_parking, rate_psf, opex_psf
  land: acreage, zoning, constr_status, price_per_acre
"""
import argparse, json, os, re, sys
from math import radians, sin, cos, asin, sqrt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# MERCURY_SURVEY_DIR overrides for testing against copies
SURVEY_DIR = Path(os.environ.get("MERCURY_SURVEY_DIR",
                                 r"C:\Users\kirkham\Documents\Amazon\Project Mercury\Surveys"))

BLD_COLS = {
    "name": 3, "address": 4, "city": 5, "state": 6, "zip": 7, "lat": 8, "lon": 9,
    "owner": 10, "total_sf": 11, "avail_sf": 12, "office_sf": 13, "land_ac": 14,
    "year_built": 15, "building_status": 16, "building_type": 17, "occupancy": 18,
    "tenancy": 19, "current_tenant": 20, "prior_use": 21, "prior_tenant": 22,
    "date_available": 23, "sprinkler": 24, "zoning": 25, "clear_min": 26,
    "clear_max": 27, "load_type": 28, "docks": 29, "grade_doors": 30,
    "col_w": 31, "col_d": 32, "auto_parking": 33, "trailer_parking": 34,
    "power": 35, "distance": 36, "sale_lease": 37, "rate_psf": 38,
    "sale_price": 39, "opex_psf": 40, "flyer": 41, "comments": 42,
}
LAND_COLS = {
    "name": 3, "address": 4, "city": 5, "state": 6, "zip": 7, "lat": 8, "lon": 9,
    "owner": 10, "acreage": 11, "zoning": 12, "constr_status": 13, "power": 14,
    "distance": 15, "sale_lease": 16, "price_per_acre": 17, "sale_price": 18,
    "flyer": 19, "comments": 20,
}

def hav(a, b, c, d):
    a, b, c, d = map(radians, (a, b, c, d))
    return 3958.8 * 2 * asin(sqrt(sin((c - a) / 2) ** 2 + cos(a) * cos(c) * sin((d - b) / 2) ** 2))

def norm(s):
    if not s:
        return ""
    s = re.sub(r"[^a-z0-9 ]", "", str(s).lower())
    for a, b in (("street", "st"), ("drive", "dr"), ("road", "rd"), ("avenue", "ave"),
                 ("boulevard", "blvd"), ("highway", "hwy"), ("lane", "ln"),
                 ("parkway", "pkwy"), ("court", "ct")):
        s = re.sub(rf"\b{a}\b", b, s)
    return re.sub(r"\s+", " ", s).strip()

def find_groups(ws, first_row):
    """Return [{pin, lat, lon, band_row, data_rows: [row#...]}] from band rows."""
    groups, cur = [], None
    for r in range(first_row, ws.max_row + 2):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        m = re.search(r"pin \((-?[\d.]+), (-?[\d.]+)\)", str(c or ""))
        if b is not None and not isinstance(b, (int, float)) and m and str(b) != "Ex.":
            cur = {"pin": str(b), "lat": float(m.group(1)), "lon": float(m.group(2)),
                   "band_row": r, "data_rows": []}
            groups.append(cur)
        elif cur is not None and isinstance(b, (int, float)):
            cur["data_rows"].append(r)
    return groups

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--msa", required=True)
    ap.add_argument("--type", required=True, choices=["building", "land"])
    ap.add_argument("--json", required=True, help="path to candidate JSON file")
    ap.add_argument("--pin", help="force a specific pin working ID")
    ap.add_argument("--overwrite", action="store_true",
                    help="on dedupe match, overwrite conflicting fields instead of reporting them")
    args = ap.parse_args()

    matches = [p for p in SURVEY_DIR.glob("* - SSD Survey.xlsx")
               if args.msa.lower() in p.name.lower()]
    if len(matches) != 1:
        sys.exit(f"ERROR: MSA '{args.msa}' matched {len(matches)} workbooks: {[p.name for p in matches]}")
    path = matches[0]

    cand = json.loads(Path(args.json).read_text(encoding="utf-8"))
    cols = BLD_COLS if args.type == "building" else LAND_COLS
    sheet = "Building Survey" if args.type == "building" else "Land Survey"
    first_row = 5 if args.type == "building" else 4

    try:
        wb = load_workbook(path)
    except PermissionError:
        sys.exit(f"ERROR: {path.name} is open in Excel — close it and re-run.")
    ws = wb[sheet]
    groups = find_groups(ws, first_row)
    if not groups:
        sys.exit(f"ERROR: no pin band rows found in {path.name} / {sheet}")

    lat, lon = cand.get("lat"), cand.get("lon")
    if args.pin:
        g = next((g for g in groups if g["pin"].lower() == args.pin.lower()), None)
        if g is None:
            sys.exit(f"ERROR: pin '{args.pin}' not in {path.name}. Pins: {[x['pin'] for x in groups]}")
    elif lat is not None and lon is not None:
        g = min(groups, key=lambda g: hav(g["lat"], g["lon"], lat, lon))
    else:
        sys.exit("ERROR: no lat/lon and no --pin — geocode the candidate first or force a pin.")

    dist = round(hav(g["lat"], g["lon"], lat, lon), 1) if lat is not None and lon is not None else None

    # criteria flags
    flags = []
    if dist is None:
        flags.append("No coordinates — verify location")
    elif dist > 5:
        flags.append(f"{dist} mi from pin — OUTSIDE 5-mi backup radius")
    elif dist > 3:
        flags.append(f"{dist} mi — inside 5-mi backup ring")
    if args.type == "building":
        av = cand.get("avail_sf")
        if isinstance(av, (int, float)) and not (75000 <= av <= 450000):
            flags.append(f"Available SF {av:,.0f} — outside 75k–450k criteria")
    else:
        ac = cand.get("acreage")
        if isinstance(ac, (int, float)) and not (7 <= ac <= 30):
            flags.append(f"{ac:g} ac — outside 7–30 ac criteria"
                         + (" (check divisibility)" if ac > 30 else ""))
    comments = ". ".join([c for c in [cand.get("comments")] + flags if c])
    cand = {**cand, "distance": dist, "comments": comments or None}

    # dedupe within the target group
    n_addr, n_name = norm(cand.get("address")), norm(cand.get("name"))
    match_row = None
    for r in g["data_rows"]:
        ex_addr = norm(ws.cell(row=r, column=cols["address"]).value)
        ex_name = norm(ws.cell(row=r, column=cols["name"]).value)
        if (n_addr and n_addr == ex_addr) or (n_name and n_name == ex_name):
            match_row = r
            break

    report = {"workbook": path.name, "sheet": sheet, "pin": g["pin"],
              "distance_mi": dist, "flags": flags}

    if match_row:
        conflicts, updated = [], []
        for k, col in cols.items():
            v = cand.get(k)
            if v is None:
                continue
            old = ws.cell(row=match_row, column=col).value
            if old in (None, "", "None"):
                ws.cell(row=match_row, column=col, value=v)
                updated.append(k)
            elif str(old) != str(v):
                if args.overwrite or k in ("distance", "comments"):
                    ws.cell(row=match_row, column=col, value=v)
                    updated.append(k)
                else:
                    conflicts.append({"field": k, "existing": old, "new": v})
        report.update(action="updated", row=match_row, fields_filled=updated, conflicts=conflicts)
    else:
        rows = g["data_rows"]
        # replace a 'None' placeholder row if that's all the group has
        if len(rows) == 1 and str(ws.cell(row=rows[0], column=3).value) == "None":
            tgt = rows[0]
            ws.cell(row=tgt, column=3, value=None)
        else:
            # insert keeping distance sort (unknown distance sorts last)
            tgt = rows[-1] + 1 if rows else g["band_row"] + 1
            if dist is not None:
                for r in rows:
                    ex = ws.cell(row=r, column=cols["distance"]).value
                    if not isinstance(ex, (int, float)) or dist < ex:
                        tgt = r
                        break
            ws.insert_rows(tgt)
        ws.cell(row=tgt, column=2, value=0)  # numeric so the renumber re-scan sees the row
        for k, col in cols.items():
            if cand.get(k) is not None:
                ws.cell(row=tgt, column=col, value=cand[k])
        # renumber the group (re-scan: rows shifted)
        groups2 = find_groups(ws, first_row)
        g2 = next(x for x in groups2 if x["pin"] == g["pin"])
        for i, r in enumerate(g2["data_rows"], 1):
            ws.cell(row=r, column=2, value=i)
        report.update(action="added", row=tgt, group_size=len(g2["data_rows"]))

    wb.save(path)
    print(json.dumps(report, indent=2, default=str))

if __name__ == "__main__":
    main()
